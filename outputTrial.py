#import xlwt
import openpyxl

Workbook = openpyxl.Workbook #creating aliases
Style = openpyxl.styles.Style #creating aliases
Alignment = openpyxl.styles.Alignment #creating aliases
Font = openpyxl.styles.Font #creating aliases

wb = Workbook()

# grab the active worksheet
ws = wb.active

ws['A1'].font = Font(bold=True)
ws['A1'].alignment = Alignment(horizontal='center')

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Inserting a data sheet to the end.
ws1 = wb.create_sheet()

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")


##x=1
##y=2
##z=3
##
##list1=[2.34,4.346,4.234]
##
##book = xlwt.Workbook(encoding="utf-8")
##
##sheet1 = book.add_sheet("Sheet 1")
##
##sheet1.write(0, 0, "Display")
##sheet1.write(1, 0, "Dominance")
##sheet1.write(2, 0, "Test")
##
##sheet1.write(0, 1, x)
##sheet1.write(1, 1, y)
##sheet1.write(2, 1, z)
##
##sheet1.write(4, 0, "Stimulus Time")
##sheet1.write(4, 1, "Reaction Time")
##
##i=4
##
##for n in list1:
##    i = i+1
##    sheet1.write(i, 0, n)
##
##book.save("trial.xls") 
