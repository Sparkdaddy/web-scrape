#! python3clear
from selenium import webdriver
from threading import Thread
import time
from datetime import datetime as dt
import timeit
import requests
import urllib.request as ur
from openpyxl import Workbook
from openpyxl.styles import Side, Alignment, Protection, Font
from bs4 import BeautifulSoup

startTime = dt.now()
##############################################################################
########################## Preparing the URLs ################################
##############################################################################

#googleURL = Static
#yahooURL = Static
#yahooStatsUrl = jsRendered
#yahooProfUrl = jsRendered

company = input("Enter in company stock symbol! ")
#Making a valid url by replacing the spaces with '+'
company = company.replace(" ", "+")
googUrl = "https://www.google.com/finance?q=" + company
yahooUrl = "https://finance.yahoo.com/quote/" + company
yahooStatsUrl = "http://finance.yahoo.com/quote/" + company + "/key-statistics?p=" + company
yahooProfUrl = "http://finance.yahoo.com/quote/" + company + "/profile?p=" + company
googPage = ur.urlopen(googUrl)
googSoup = BeautifulSoup(googPage, "html.parser")
yahooPage = ur.urlopen(yahooUrl)
yahooSoup = BeautifulSoup(yahooPage, "html.parser")

driver = webdriver.Chrome("/Users/Sparky/Downloads/chromedriver")
driver.get(yahooStatsUrl)
time.sleep(5)
yahooStatsPage = driver.page_source
yahooStatsSoup = BeautifulSoup(yahooStatsPage, "html.parser")
driver.get(yahooProfUrl)
time.sleep(5)
yahooProfPage = driver.page_source
yahooProfSoup = BeautifulSoup(yahooProfPage, "html.parser")
driver.quit()


##############################################################################
################################## Output ####################################
##############################################################################

wb = Workbook() #creating the workbook in excel
ws = wb.active #grabbing the active worksheet
ws.title = "Stock Information" #changing the sheet title
ws1 = wb.create_sheet() #creating a new sheet
ws1.title = "Rating"

ws['A1'] = 'Daily Information' #inserting info into sheet
ws['E1'] = 'Historical Stock Information'
ws['I1'] = 'Shareholders'
ws['A13'] = 'Statistical Overlook'
ws['E13'] = 'Sentiment/trend'
ws['A22'] = 'Management'
ws['E20'] = '??Comparison trend line of stock to these??' #do you want these to be chart comparisons? Yes

ws['A3'] = 'Price'
ws['A4'] = 'Volume'
ws['A5'] = 'Average Volume (3m)'
ws['A6'] = 'Market Cap'
ws['A7'] = 'Ask'
ws['A8'] = 'Earnings Date'
ws['A9'] = 'Category'
ws['A10'] = 'Lagger/Leader'

ws['E3'] = 'Historical PE Ratio'
ws['E4'] = 'Historical EPS'
ws['E5'] = 'Premium/Discount'
ws['E6'] = 'Historical Premium/Discount'
ws['E7'] = 'Dividend Payout Ratio/Trend'
ws['E8'] = 'Number of employees historically'
ws['E9'] = '??Trend Line Showing price in Comparison to historical eps/pe??'

ws['I3'] = 'Outstanding shares'
ws['I4'] = 'Institutionally Held'
ws['I5'] = 'Biggest Shareholders'
ws['I6'] = 'Held by Insiders'
ws['I7'] = 'Industry averages for these'

ws['A14'] = 'Standard deviation (annual)'
ws['A15'] = 'Projected 5 year growth'
ws['A16'] = 'number of employees'
ws['A17'] = 'Standard deviation (3 months)'
ws['A18'] = '50 day moving average'
ws['A19'] = '200 day moving average'

ws['A20'] = 'Analyst Sentiment'

ws['A23'] = 'Latest happenings with management'

ws['E21'] = 'Industry'
ws['E22'] = 'Dow Jones'
ws['E23'] = 'S&P'
ws['E24'] = 'Nasdaq'

###########################Styling###################

def Styling(headerCells):
    headerCells.font = Font(bold=True)
    headerCells.alignment = Alignment(horizontal='center')
Styling(ws['A1'])
Styling(ws['E1'])
Styling(ws['I1'])
Styling(ws['A13'])
Styling(ws['E13'])
Styling(ws['A20'])
Styling(ws['E20'])
ws.column_dimensions["A"].width = 22.0
ws.column_dimensions["E"].width = 22.0
ws.column_dimensions["I"].width = 22.0
ws.merge_cells('A1:B1')
ws.merge_cells('E1:G1')
ws.merge_cells('I1:J1')
ws.merge_cells('A13:B13')
ws.merge_cells('E13:F13')
ws.merge_cells('E20:F20')


##############################################################################
############################ Scraping daily info #############################
##############################################################################
def daily_info():
    d = {}
    var = ['range','range_52week', 'vol_and_avg', 'market_cap', 'pe_ratio', 'eps', 'beta']

    for i in var:
        #Searches for the table data containing the key "data-snapfield = range"
        #May be quicker to change from searches to manipulation through next sibling
        val = googSoup.find('td',{'data-snapfield':i})
        #Temporarily moving line where the actual data is. Necessary as the
        #data is repetitive with similar keys.
        temp = val.next_sibling.next_sibling.get_text()

        #rstrip -- removing the endline character
        d[i] = temp.rstrip('\n')

    ws['B3'] = d['range']
    ws['B4'] = d['range_52week']
    ws['B5'] = d['vol_and_avg']
    ws['B6'] = d['market_cap']
    ws['B7'] = d['pe_ratio']
    ws['B8'] = d['eps']
    ws['B9'] = d['beta']

try:
    daily_info()
except AttributeError:
    print("Oh no! The daily data was not found!")
    print("\nHere is the URL") + googUrl
##############################################################################
############################ Scraping stats info #############################
##############################################################################

# val = yahooStatsSoup.find(string = "200-Day Moving Average") seventh item
val = yahooStatsSoup.find_all('table' ,{'class' : "table-qsp-stats Mt(10px)"})
count = 0
temp2
for i in val:
    count += 1
    if count = 8:
        temp2 = i
day50Average = val.children.children.next_sibling.next_sibling.next_sibling.next_sibling
day200Average = day50Average.next_sibling
print(temp)
print(val)
ws['B19'] = day200Average.get_text()
ws['B18'] = day50Average.get_text()
#print(yahooStatsSoup.prettify())
#The data is generated using reactjs. Thus it doesn't exist using this parser.
# print(val)
# try:
#     temp = val.get_text()
#     print (temp)
# except AttributeError:
#     print("Oh no. The value was not found")
#     print("Here is the URL:")
#     print(yahooStatsUrl)
wb.save(company+str(dt.now())+".xlsx")
stopTime = dt.now()
totalTime = startTime - stopTime
print("The total time taken is: " + str(float(totalTime.total_seconds())))
